import xlrd
import xlwt
import openpyxl


def processing_excel():
    workbook = openpyxl.load_workbook(r'/home/hirain777/hanzzi_learn_carla/data/Highway_dynamic.xlsx')
    target_sheet = workbook.get_sheet_by_name("Raw_data")
    row_num = target_sheet.max_row
    col_num = target_sheet.max_column
    print (col_num)
    fps = 22
    data_num = int(row_num) / fps

    # print (data_num)
    all_data = []
    for col_index in range (1, col_num + 1):
        output = []
        for data_num_index in range (0,data_num):
            sum = 0.0
            for row_index in range (1,fps+1):
                if target_sheet.cell(row = data_num_index * fps + row_index + 1,column = col_index ).value is None:
                    current_value =  0.0
                else:
                    current_value = target_sheet.cell(row = data_num_index * fps + row_index + 1, column = col_index).value
                sum = sum + current_value
            output.append(sum/float(fps))
        all_data.append(output)
    
    workbook.create_sheet(title='Postprocessing')
    write_sheet = workbook.get_sheet_by_name(workbook.sheetnames[-1])
    
    for col_index in range (1, col_num + 1):
        write_sheet.cell(row = 1, column = col_index).value =  target_sheet.cell(row = 1, column = col_index).value

    for col_index in range (1, col_num+1):
        for row_index in range (1, data_num):
            write_sheet.cell(row = row_index + 1, column = col_index).value = all_data[col_index-1][row_index-1]
    
    workbook.save(r'/home/hirain777/hanzzi_learn_carla/data/Highway_dynamic.xlsx')

if __name__ == '__main__':
    processing_excel()   